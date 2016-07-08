from django.contrib.staticfiles.templatetags.staticfiles import static
from django.apps import AppConfig
from django.core import serializers
from django.db import OperationalError

import openpyxl

Attr = {
	'pk': 0,
	'ecole': 1,
	'type': 2,
	'ville': 3,
	'adresse': 4,
	'programmes': 5,
	'url': 6
}

def load_data_from_excel(Ecole) :
	app_name = 'carte_interactive'
	data_url = static('carte_interactive/data/data.xlsx')
	ecole_data = openpyxl.load_workbook(app_name + data_url)
	sheet = ecole_data.get_sheet_by_name('data')
	ecoles = sheet.rows
	try:
		for row in ecoles:
			# create Ecole objects with content of Excel file
			Ecole.objects.update_or_create(
				pk=row[Attr['pk']].value,
				defaults= {
					'nom': row[Attr['ecole']].value,
					'type': row[Attr['type']].value,
					'ville': row[Attr['ville']].value,
					'programmes': row[Attr['programmes']].value,
					'adresse': row[Attr['adresse']].value,
					'url': row[Attr['url']].value
				}
			)

		# fill json file with Ecole data, for use with Google Javascript API
		json_data = serializers.serialize('json', Ecole.objects.all())
		json_data_url = static('carte_interactive/json/data.json')
		json_data_file = open(app_name + json_data_url, 'w')
		json_data_file.write(json_data)
		json_data_file.close()
	except OperationalError:
		pass

# rewrite excel file with other excel file content
def rewriteExcel(data_file_path, uploaded_file):
	uploaded_wb = openpyxl.load_workbook(uploaded_file)
	uploaded_sheet_name = uploaded_wb.get_sheet_names()[0]
	uploaded_sheet = uploaded_wb.get_sheet_by_name(uploaded_sheet_name)

	data_wb = openpyxl.load_workbook(data_file_path)
	data_sheet = data_wb.get_sheet_by_name('data')

	# wipe data file
	data_ecoles = data_sheet.rows
	try:
		for row in data_ecoles:
			for cell in row:
				cell.value = ""

		data_wb.save(data_file_path)
	except OperationalError:
		pass

	# rewrite data file
	uploaded_ecoles = uploaded_sheet.rows
	print(uploaded_ecoles[1][Attr['pk']].value)
	row_count = 1;
	try:
		for row in uploaded_ecoles[1:]:
			# write row into data.xlsx file
			data_sheet.cell(row=row_count, column = Attr['pk']+1).value = row[Attr['pk']].value
			data_sheet.cell(row=row_count, column = Attr['ecole']+1).value = row[Attr['ecole']].value
			data_sheet.cell(row=row_count, column = Attr['type']+1).value = row[Attr['type']].value
			data_sheet.cell(row=row_count, column = Attr['ville']+1).value = row[Attr['ville']].value
			data_sheet.cell(row=row_count, column = Attr['programmes']+1).value = row[Attr['programmes']].value
			data_sheet.cell(row=row_count, column = Attr['adresse']+1).value = row[Attr['adresse']].value
			data_sheet.cell(row=row_count, column = Attr['url']+1).value = row[Attr['url']].value
			row_count += 1

		data_wb.save(data_file_path)
	except OperationalError:
		pass

